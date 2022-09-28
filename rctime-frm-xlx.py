"""
It's collecting the B5K's re-convergent data, based on below rules:
1. record the period which 'Fix Type' in xlx from RTX(9) to none RTX as RTX Lost Period, 
    or 'GNSS State' in xlx from useing GNSS (0) to not using GNSS (5) as GNSS lost period.
1.1 have to be manually selected (x_trig, c_trig), I'm not going to make it automatically or give all avaiable
    no one care about this code, and I'm lazy.

2. when acitve "average" value for re-convergent calculation (accy_list[0] == 0)
2.1 when it goes back to RTX from lost, skip skip_num ahead of data from the previous RTX fix position, for stable value
2.2 then start collect continuous avg_num numbers of average "Pos Accy (m)" value (accy_1) as comparision base number
2.3 check continuous avg_num numbers of average "Pos Accy (m)" value (accy_2) beginning with RTX recovered point
2.4 once accy_2 <= accy_1, consider this preriod as re-convergent time

3. when acitve "list" value for re-convergent calculation (accy_list[0] != 0)
3.1 when it goes back to RTX from lost, only compare the "Pos Accy (m)" value with the numbers in accy_list
"""

xver        = '0.10'
import      os
import      sys
import      openpyxl

#------change below code based on your requirement------

#xlx_path    = 'C:\\Work\\Tools\\rctime_from_xlx\\'
xlx_path    = '/Users/Hawk/Downloads/'
xlx_name    = 'sv_example'
xlx_tail    = '.xlsx'
output_tail = '.csv'

# RTX lost period threshold (s), 0 means no threshold
L_threshold = 0

# it has to be only 3, and has to be (left one > right one)
# if accy_list[0] == 0, it means use average value instead of list value
accy_list   = [0.1, 0.05, 0.02]

#2.1 when it goes back to RTX from lost, skip skip_num ahead of data from the previous RTX fix position, for stable value
skip_num    = 5#10

#2.2 then start collect continuous avg_num numbers of average "Pos Accy (m)" value (accy_1) as comparision base number
#2.3 check continuous avg_num numbers of average "Pos Accy (m)" value (accy_2) beginning with RTX recovered point
avg_num     = 8#10

# show more information in infoX
for_dbg     = 0

#------DO NOT change below code if you don't know how to do so------

xlx_wb = openpyxl.load_workbook(xlx_path+xlx_name+xlx_tail)
xlx_sht = xlx_wb[xlx_wb.sheetnames[0]]

x_gpstime   = 3
x_time      = 7
#x_trig      = 18 #use 'Fix Type' for trigger RTX lost period
x_trig      = 37 #use 'GNSS State' to trigger the GNSS lost period
x_lon       = 9
x_lat       = 10
x_posacc    = 29
x_gps_max   = 86400#24 hours in seconds
x_gps       = 0
x_gps_l     = [0, 0, 0]
c_count     = 0
#c_trig      = 9 #use 'Fix Type' for trigger RTX lost period
c_trig      = 0 #use 'GNSS State' to trigger the GNSS lost period
c_flag      = 0
c_acc_1     = 0
c_acc_2     = 0
c_row_fix   = 0

c_msg = '\nVersion: ' + xver
print(c_msg)
c_msg += '\n'
c_msg += '\n info1 - fix mode lost again'
c_msg += '\n info2 - previous index smaller than skip+avgnum'
c_msg += '\n info3 - previous fixed average item not in fix mode'
c_msg += '\n info4 - fixed average item not in fix mode'
c_msg += '\n info5 - RTX lost period too short'
c_msg += '\n info6 - file end'
c_msg += '\n'
c_msg += '\nIndex,StartLon,StartLat,StartTime,EndLon,EndLat,EndTime,RTXLostPeriod(s),'

if accy_list[0] == 0:
    c_msg += 'StartAccy,EndAccy,ReCnvtPeriod(s)'
    output_file = xlx_path+xlx_name+'_avg'+output_tail
else:
    if accy_list[0] <= accy_list[1] or accy_list[1] <= accy_list[2]:
        print('\nList value error: ' + str(accy_list[0]) + ' ' + str(accy_list[1]) + ' ' + str(accy_list[2]))
        sys.exit()
    c_msg += str(accy_list[0])+'M(s),'+str(accy_list[1])+'M(s),'+str(accy_list[2])+'M(s)'
    output_file = xlx_path+xlx_name+'_list'+output_tail

# x: Open a file for exclusive creation. If the file already exists, the operation fails.
# w: Open a file for writing. Creates a new file if it does not exist or truncates the file if it exists.
with open(output_file, 'w') as c_log:
    c_log.write(c_msg)
    for row_1 in range(2, xlx_sht.max_row+1):
        if c_trig != xlx_sht.cell(row_1, x_trig).value:
            # Never goes into fix yet
            if c_flag == 0: 
                continue
            # Fix lost start
            elif c_flag == 1: 
                if x_gps == 0:
                    c_count += 1
                    x_gps = xlx_sht.cell(row_1, x_gpstime).value
                    c_msg = '\n'
                    c_msg += str(c_count).zfill(3) # Index
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_lon).value) # StartLon
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_lat).value) # StartLat
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_time).value) # StartTime
                    # Keep item index for later calculation
                    c_row_fix = row_1
            # Fix lost again before finished the convergent calculate
            else:
                c_flag = 1 # Fix lost start
                x_gps = 0 
                x_gps_l[0] = 0 
                x_gps_l[1] = 0 
                x_gps_l[2] = 0
                if for_dbg:
                    c_msg += ',info1(' + str(row_1) + ')'
                else:
                    c_msg += ',info1'
                c_log.write(str(c_msg))
                c_msg = 0
                continue
        else:# Goes into fix
            if c_flag == 0: 
                c_flag = 1 # Fix lost start
                continue
            elif c_flag == 1: # Fix lost end
                if x_gps != 0:
                    if xlx_sht.cell(row_1, x_gpstime).value < x_gps:
                        x_gps = x_gps_max - x_gps
                        x_gps += xlx_sht.cell(row_1, x_gpstime).value
                    else:
                        x_gps = xlx_sht.cell(row_1, x_gpstime).value - x_gps
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_lon).value) # EndLon
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_lat).value) # EndLat
                    c_msg += ','
                    c_msg += str(xlx_sht.cell(row_1, x_time).value) # EndTime
                    c_msg += ','
                    c_msg += str(format(x_gps,'.2f')) # RTXLostPeriod(s)


                    if L_threshold and x_gps <= L_threshold:
                        if for_dbg:
                            c_msg += ',info5(' + str(format(x_gps,'.2f')) + '<=' + str(format(L_threshold,'.2f')) + ')'
                        else:
                            c_msg += ',info5'
                        c_log.write(c_msg)
                        c_msg = 0
                        x_gps = 0 
                        x_gps_l[0] = 0 
                        x_gps_l[1] = 0 
                        x_gps_l[2] = 0
                        c_flag = 0
                        continue

                    # Start the reconvergence calculation
                    c_acc_1 = 0
                    c_flag = 2 
                    # Keep it for later calculation
                    x_gps = xlx_sht.cell(row_1, x_gpstime).value
                    if accy_list[0] != 0:
                        x_gps_l[0] = xlx_sht.cell(row_1, x_gpstime).value 
                        x_gps_l[1] = xlx_sht.cell(row_1, x_gpstime).value 
                        x_gps_l[2] = xlx_sht.cell(row_1, x_gpstime).value 
                        continue
                    
                    # Average of avg_num numbers of data after skip skip_num numbers of data 
                    if c_row_fix < (skip_num + avg_num):
                        c_flag = 0
                        x_gps = 0
                        if for_dbg:
                            c_msg += ',info2(' + str(c_row_fix) + ')'
                        else:
                            c_msg += ',info2'
                        c_log.write(c_msg)
                        c_msg = 0
                        continue
                    for row_2 in range(c_row_fix - skip_num - avg_num, c_row_fix - skip_num):
                        # make sure each item still in fix mode
                        if c_trig == xlx_sht.cell(row_2, x_trig).value:
                            c_acc_1 +=  xlx_sht.cell(row_2, x_posacc).value
                        else:
                            c_flag = 1 # Fix lost start
                            x_gps = 0
                            if for_dbg:
                                c_msg += ',info3(' + str(row_2) + ')'
                            else:
                                c_msg += ',info3'
                            c_log.write(c_msg)
                            c_msg = 0
                            break
                    if c_msg == 0:
                        continue
                    c_acc_1 /= avg_num
                    c_msg += ','
                    c_msg += str(format(c_acc_1,'.2f')) # StartAccy
            elif c_flag == 2: # Fix back, start reconvergence calculation
                if accy_list[0] != 0:
                    for i in range(3):
                        if x_gps_l[i] != 0:
                            if accy_list[i] >= xlx_sht.cell(row_1, x_posacc).value:
                                if xlx_sht.cell(row_1, x_gpstime).value < x_gps_l[i]:
                                     x_gps_l[i] = x_gps_max - x_gps_l[i]
                                     x_gps_l[i] += xlx_sht.cell(row_1, x_gpstime).value
                                else:
                                     x_gps_l[i] = xlx_sht.cell(row_1, x_gpstime).value - x_gps_l[i]
                                c_msg += ','
                                c_msg += str(format(x_gps_l[i],'.2f')) 
                                x_gps_l[i]  = 0
                    if x_gps_l[0] !=0 or x_gps_l[1] !=0 or x_gps_l[2] != 0:
                        continue
                else:
                    # Start to get average of avg_num number of data
                    if c_acc_1 >= xlx_sht.cell(row_1, x_posacc).value:
                        # Average of avg_num numbers of data
                        c_acc_2 = 0
                        for row_2 in range(row_1, row_1+avg_num):
                            # make sure each item still in fix mode
                            if c_trig == xlx_sht.cell(row_2, x_trig).value:
                                c_acc_2 +=  xlx_sht.cell(row_2, x_posacc).value
                            else:
                                c_flag = 1 # Fix lost start
                                x_gps = 0
                                if for_dbg:
                                    c_msg += ',info4(' + str(row_2) + ')'
                                else:
                                    c_msg += ',info4'
                                c_log.write(c_msg)
                                c_msg = 0
                                break
                        if c_msg == 0:
                            continue
                        c_acc_2 /= avg_num
                        # Check if the average lower than expected
                        if c_acc_1 < c_acc_2:
                            continue
                        if xlx_sht.cell(row_1, x_gpstime).value < x_gps:
                            x_gps = x_gps_max - x_gps
                            x_gps += xlx_sht.cell(row_1, x_gpstime).value
                        else:
                            x_gps = xlx_sht.cell(row_1, x_gpstime).value - x_gps

                        c_msg += ','
                        c_msg += str(format(c_acc_2,'.2f')) # EndAccy
                        c_msg += ',' 
                        c_msg += str(format(x_gps,'.2f')) # ReCnvtPeriod(s)
                    else:
                        continue
                c_log.write(c_msg)
                c_msg = 0
                x_gps = 0 
                x_gps_l[0] = 0 
                x_gps_l[1] = 0 
                x_gps_l[2] = 0
                c_flag = 0
    if c_msg: # collect the last line
        if for_dbg:
            c_msg += ',info6(' + str(row_1) + ')'
        else:
            c_msg += ',info6'
        c_log.write(c_msg)
    c_msg = '\n\nProcess ' + str(row_1 - 1) + ' lines\n'
    c_msg += '\nLog path:' + output_file
    c_log.write(c_msg)
print(c_msg)
